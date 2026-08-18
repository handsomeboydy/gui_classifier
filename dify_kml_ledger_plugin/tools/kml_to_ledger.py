"""Dify tool entrypoint for KML ledger conversion."""

from __future__ import annotations

from collections.abc import Generator
from dataclasses import dataclass
from hashlib import sha256
from html import unescape
from io import BytesIO
import os
from pathlib import Path
import re
from typing import Any, Optional
import xml.etree.ElementTree as ET

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


class ConversionError(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class Tower:
    number: int
    longitude: float
    latitude: float
    name: str


@dataclass(frozen=True)
class ParsedIdentity:
    source_line_name: str
    line_full_name: str
    voltage_level: str
    line_names: tuple[str, ...]
    circuit_type: str
    warnings: tuple[str, ...] = ()
    manual_required: bool = False
    manual_reason: str = ""


_VOLTAGE_PATTERN = re.compile(r"(?i)(?<![A-Za-z0-9])([+-]?\s*\d+(?:\.\d+)?)\s*(?:kV|千伏)(?![A-Za-z])")
_RANGE_SUFFIX_PATTERN = re.compile(
    r"(?:\s*[_-]?\s*)(?:N\s*)?\d+\s*[-~—至]\s*(?:N\s*)?\d+\s*$",
    re.I,
)


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1].lower()


def normalize_voltage(value: object) -> str:
    text = str(value or "").strip().replace("千伏", "kV")
    text = re.sub(r"\s+", "", text)
    return text[:-2] if text.lower().endswith("kv") else text


def parse_identity(filename: str, parameters: dict[str, Any]) -> ParsedIdentity:
    source_file_name = os.path.basename(str(filename or "").strip())
    stem = Path(source_file_name).stem.strip()
    if not stem:
        return ParsedIdentity("", "", "", (), "", manual_required=True, manual_reason="KML 文件名为空")
    normalized = _RANGE_SUFFIX_PATTERN.sub("", stem).strip(" _-")
    voltage_match = _VOLTAGE_PATTERN.search(normalized)
    parsed_voltage = normalize_voltage(voltage_match.group(1)) if voltage_match else ""
    full_name = normalized
    if voltage_match:
        full_name = (normalized[:voltage_match.start()] + normalized[voltage_match.end():]).strip(" _-")
    full_name = re.sub(r"\s+", "", full_name)

    voltage = normalize_voltage(parameters.get("voltage_level")) or parsed_voltage
    line_1 = str(parameters.get("line_name_1") or "").strip()
    line_2 = str(parameters.get("line_name_2") or "").strip()
    circuit = str(parameters.get("circuit_type") or "").strip().lower()
    if line_1 or line_2 or parameters.get("voltage_level") or circuit:
        if not voltage:
            return ParsedIdentity(full_name, full_name, "", (), "", manual_required=True, manual_reason="请补充电压等级")
        if line_1 and (line_2 or circuit in {"double", "双回", "甲乙"}):
            if not line_2:
                return ParsedIdentity(full_name, full_name, voltage, (), "双回", manual_required=True, manual_reason="双回线路必须同时填写线路1和线路2名称")
            return ParsedIdentity(full_name, full_name, voltage, (line_1, line_2), "双回")
        if line_1:
            return ParsedIdentity(full_name, full_name, voltage, (line_1,), "单回")
        return ParsedIdentity(full_name, full_name, voltage, (), "", manual_required=True, manual_reason="手动输入缺少线路1名称")

    warnings = []
    if not parsed_voltage:
        warnings.append("未从文件名识别电压等级")
    double_match = re.search(r"(?P<base>.*?)(?:甲乙|甲、乙|甲/乙)线$", full_name)
    if double_match:
        base = double_match.group("base")
        names = ((base + "甲线") if base else "甲线", (base + "乙线") if base else "乙线")
        circuit_type = "双回"
    elif re.search(r"(?:甲|乙)线$", full_name):
        names = (full_name,)
        circuit_type = "单回"
    elif full_name:
        names = (full_name,)
        circuit_type = "单回"
    else:
        names = ()
        circuit_type = ""
    reasons = []
    if not parsed_voltage:
        reasons.append("无法识别电压等级")
    if not names:
        reasons.append("无法识别线路名称")
    return ParsedIdentity(
        full_name,
        full_name,
        parsed_voltage,
        names,
        circuit_type,
        tuple(warnings),
        manual_required=bool(reasons),
        manual_reason="；".join(reasons),
    )


def child_text(element: ET.Element, tag_name: str) -> str:
    for child in element.iter():
        if local_name(child.tag) == tag_name:
            return "".join(child.itertext()).strip()
    return ""


def tower_number(name: str, description: str) -> Optional[int]:
    text_values = (re.sub(r"<[^>]*>", " ", unescape(description or "")), name)
    patterns = (
        re.compile(r"(?i)(?:^|[^A-Z0-9])N\s*0*(\d+)(?:\D|$)"),
        re.compile(r"(?:杆塔编号|杆塔号|塔号|塔位号|塔位)\s*[:：#№]?\s*(\d+)", re.I),
        re.compile(r"(\d+)\s*号\s*(?:杆塔|塔)", re.I),
        re.compile(r"#\s*(\d+)"),
    )
    for value in text_values:
        for pattern in patterns:
            match = pattern.search(value)
            if match and int(match.group(1)) > 0:
                return int(match.group(1))
        if re.fullmatch(r"\s*\d+\s*", value or "") and int(value) > 0:
            return int(value)
    return None


def parse_kml(raw: bytes) -> tuple[list[Tower], list[str]]:
    if raw.startswith(b"\xff\xfe"):
        text = raw[2:].decode("utf-16le")
    elif raw.startswith(b"\xfe\xff"):
        text = raw[2:].decode("utf-16be")
    else:
        text = raw.decode("utf-8-sig")
    try:
        root = ET.fromstring(text)
    except ET.ParseError as exc:
        raise ConversionError("KML_XML_INVALID", f"KML XML 解析失败: {exc}") from exc

    placemarks = [element for element in root.iter() if local_name(element.tag) == "placemark"]
    candidates = []
    warnings = []
    for index, placemark in enumerate(placemarks, 1):
        point = next((item for item in placemark.iter() if local_name(item.tag) == "point"), None)
        if point is None:
            continue
        name = child_text(placemark, "name")
        coordinate_text = child_text(point, "coordinates").strip()
        coordinate = coordinate_text.split()[0] if coordinate_text else ""
        parts = coordinate.split(",")
        if len(parts) < 2:
            raise ConversionError("COORDINATE_INVALID", f"第 {index} 个杆塔点坐标无效: {name or '未命名'}")
        try:
            longitude, latitude = float(parts[0]), float(parts[1])
        except ValueError as exc:
            raise ConversionError("COORDINATE_INVALID", f"第 {index} 个杆塔点坐标不是数值: {name or '未命名'}") from exc
        if not (-180 <= longitude <= 180 and -90 <= latitude <= 90):
            raise ConversionError("COORDINATE_INVALID", f"第 {index} 个杆塔点坐标越界: {longitude}, {latitude}")
        candidates.append([tower_number(name, child_text(placemark, "description")), longitude, latitude, name])
    if not candidates:
        raise ConversionError("NO_TOWER_POINTS", "KML 中未找到带有效 Point 坐标的杆塔")

    used = {item[0] for item in candidates if item[0] is not None}
    next_number = 1
    towers = []
    for number, longitude, latitude, name in candidates:
        if number is None:
            while next_number in used:
                next_number += 1
            number = next_number
            used.add(number)
            next_number += 1
            warnings.append(f"杆塔 {name or '未命名'} 未识别编号，已按顺序使用 {number}")
        towers.append(Tower(number, longitude, latitude, name))
    numbers = [tower.number for tower in towers]
    duplicate = sorted({number for number in numbers if numbers.count(number) > 1})
    if duplicate:
        raise ConversionError("TOWER_NUMBER_DUPLICATED", f"存在重复杆塔编号: {', '.join(map(str, duplicate))}")
    towers.sort(key=lambda tower: tower.number)
    return towers, warnings


def make_xlsx(
    line_name: str,
    towers: list[Tower],
    *,
    line_full_name: str,
    voltage_level: str,
    circuit_type: str,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "经纬度台账"
    sheet.freeze_panes = "A2"
    sheet.append(["线路名称", "杆塔编号", "经度", "纬度", "线路全称", "电压等级(kV)", "线路类型"])
    for tower in towers:
        sheet.append([
            line_name,
            tower.number,
            tower.longitude,
            tower.latitude,
            line_full_name,
            voltage_level,
            circuit_type,
        ])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in list(sheet["C"])[1:] + list(sheet["D"])[1:]:
        cell.number_format = "0.000000"
    for column, width in {"A": 22, "B": 14, "C": 14, "D": 14, "E": 24, "F": 14, "G": 12}.items():
        sheet.column_dimensions[column].width = width
    table = Table(displayName="TowerLedger", ref=f"A1:G{len(towers) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class KmlToLedgerTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        file = tool_parameters.get("file")
        if not file:
            raise ValueError("file is required")
        filename = os.path.basename(file.filename or "")
        if not filename.lower().endswith(".kml"):
            raise ConversionError("INVALID_FILE_TYPE", "只支持 .kml 文件")
        raw = file.blob
        if not raw:
            raise ConversionError("KML_DECODE_FAILED", "KML 文件内容为空")
        towers, warnings = parse_kml(raw)
        identity = parse_identity(filename, tool_parameters)
        all_warnings = list(identity.warnings) + list(warnings)
        if identity.manual_required:
            yield self.create_variable_message("status", "manual_required")
            yield self.create_variable_message("manual_required", True)
            yield self.create_variable_message("manual_reason", identity.manual_reason)
            yield self.create_variable_message("file_names", [])
            yield self.create_variable_message("line_names", list(identity.line_names))
            yield self.create_variable_message("voltage_level", identity.voltage_level)
            yield self.create_variable_message("tower_count", len(towers))
            yield self.create_variable_message("warnings", all_warnings)
            return

        file_names = []
        line_names = []
        sha256s = []
        for line_name in identity.line_names:
            output = make_xlsx(
                line_name,
                towers,
                line_full_name=identity.line_full_name,
                voltage_level=identity.voltage_level,
                circuit_type=identity.circuit_type,
            )
            output_name = f"{line_name}经纬度台账.xlsx"
            file_names.append(output_name)
            line_names.append(line_name)
            sha256s.append(sha256(output).hexdigest())
            yield self.create_blob_message(
                blob=output,
                meta={
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "filename": output_name,
                },
            )
        yield self.create_variable_message("status", "succeeded")
        yield self.create_variable_message("manual_required", False)
        yield self.create_variable_message("manual_reason", "")
        # Do not emit a variable named ``files`` here.  The XLSX blobs above
        # are automatically exposed by Dify as the tool node's ``files``
        # (array[file]) output.  Emitting the same name as an array of
        # filename strings creates a type conflict in the workflow End node.
        yield self.create_variable_message("file_names", file_names)
        yield self.create_variable_message("line_names", line_names)
        yield self.create_variable_message("voltage_level", identity.voltage_level)
        yield self.create_variable_message("line_full_name", identity.line_full_name)
        yield self.create_variable_message("circuit_type", identity.circuit_type)
        yield self.create_variable_message("tower_count", len(towers))
        yield self.create_variable_message("warnings", all_warnings)
        yield self.create_variable_message("sha256", sha256s[0] if len(sha256s) == 1 else "")
        yield self.create_variable_message("sha256s", sha256s)
