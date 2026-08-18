"""Deterministic KML -> XLSX ledger conversion.

The converter is deliberately independent from Tkinter and Dify.  It accepts
bytes or a local file, validates tower points and coordinates, then produces
the exact four-column workbook consumed by the classifier.
"""

from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from html import unescape
from io import BytesIO
import os
from pathlib import Path
import re
from typing import Iterable, Optional
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


SUPPORTED_EXTENSION = ".kml"
MAX_INPUT_BYTES = 15 * 1024 * 1024
HEADERS = ("线路名称", "杆塔编号", "经度", "纬度")
METADATA_HEADERS = ("线路全称", "电压等级(kV)", "线路类型")


class KmlConversionError(ValueError):
    """A user-actionable conversion failure with a stable error code."""

    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code
        self.message = message


@dataclass(frozen=True)
class TowerPoint:
    tower: int
    longitude: float
    latitude: float
    source_name: str
    used_fallback_number: bool = False


@dataclass(frozen=True)
class ConversionResult:
    line_name: str
    file_name: str
    tower_count: int
    warnings: tuple[str, ...]
    xlsx_bytes: bytes
    sha256: str
    line_full_name: str = ""
    voltage_level: str = ""
    circuit_type: str = "单回"

    @property
    def warning_count(self) -> int:
        return len(self.warnings)


@dataclass(frozen=True)
class ParsedLineIdentity:
    """Normalized line identity derived from a KML filename or manual overrides."""

    source_line_name: str
    line_full_name: str
    voltage_level: str
    line_names: tuple[str, ...]
    circuit_type: str
    warnings: tuple[str, ...] = ()
    manual_required: bool = False
    manual_reason: str = ""


@dataclass(frozen=True)
class LedgerArtifact:
    file_name: str
    line_name: str
    tower_count: int
    warnings: tuple[str, ...]
    xlsx_bytes: bytes
    sha256: str
    line_full_name: str = ""
    voltage_level: str = ""
    circuit_type: str = "单回"


@dataclass(frozen=True)
class BatchConversionResult:
    source_file_name: str
    source_line_name: str
    line_full_name: str
    voltage_level: str
    circuit_type: str
    artifacts: tuple[LedgerArtifact, ...]
    warnings: tuple[str, ...]
    manual_required: bool = False
    manual_reason: str = ""


_TAG_NUMBER_PATTERNS = (
    re.compile(r"(?i)(?:^|[^A-Z0-9])N\s*0*(\d+)(?:\D|$)"),
    re.compile(r"(?:杆塔编号|杆塔号|塔号|塔位号|塔位)\s*[:：#№]?\s*(\d+)", re.I),
    re.compile(r"(\d+)\s*号\s*(?:杆塔|塔)", re.I),
    re.compile(r"#\s*(\d+)"),
)
_PURE_NUMBER = re.compile(r"^\s*(\d+)\s*$")
_VOLTAGE_PATTERN = re.compile(r"(?i)(?<![A-Za-z0-9])([+-]?\s*\d+(?:\.\d+)?)\s*(?:kV|千伏)(?![A-Za-z])")
_RANGE_SUFFIX_PATTERN = re.compile(
    r"(?:\s*[_-]?\s*)(?:N\s*)?\d+\s*[-~—至]\s*(?:N\s*)?\d+\s*$",
    re.I,
)


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1].lower()


def _normalize_voltage(value: object) -> str:
    text = str(value or "").strip().replace("千伏", "kV")
    text = re.sub(r"\s+", "", text)
    if text.lower().endswith("kv"):
        text = text[:-2]
    return text


def parse_line_identity(filename: str, overrides: Optional[dict[str, object]] = None) -> ParsedLineIdentity:
    """Parse voltage/circuit names from an inconsistent KML filename.

    Standard examples include ``220kV砚利甲乙线N1-N120.kml`` and
    ``220kV利榄甲线  N1-N48.kml``.  Manual values take precedence when
    supplied by the desktop client.
    """

    overrides = overrides or {}
    source_file_name = os.path.basename(str(filename or "").strip())
    stem = Path(source_file_name).stem.strip()
    if not stem:
        return ParsedLineIdentity("", "", "", (), "", manual_required=True, manual_reason="KML 文件名为空")

    normalized = _RANGE_SUFFIX_PATTERN.sub("", stem).strip(" _-")
    voltage_match = _VOLTAGE_PATTERN.search(normalized)
    parsed_voltage = _normalize_voltage(voltage_match.group(1)) if voltage_match else ""
    parsed_full_name = normalized
    if voltage_match:
        parsed_full_name = (normalized[:voltage_match.start()] + normalized[voltage_match.end():]).strip(" _-")
    parsed_full_name = re.sub(r"\s+", "", parsed_full_name)

    manual_voltage = _normalize_voltage(overrides.get("voltage_level"))
    voltage = manual_voltage or parsed_voltage
    manual_line_1 = str(overrides.get("line_name_1") or "").strip()
    manual_line_2 = str(overrides.get("line_name_2") or "").strip()
    circuit_override = str(overrides.get("circuit_type") or "").strip().lower()

    warnings: list[str] = []
    if manual_line_1 or manual_line_2 or manual_voltage or circuit_override:
        if manual_line_1 and (manual_line_2 or circuit_override in {"double", "双回", "甲乙"}):
            if not manual_line_2:
                return ParsedLineIdentity(
                    parsed_full_name,
                    parsed_full_name,
                    voltage,
                    (),
                    "双回",
                    manual_required=True,
                    manual_reason="双回线路必须同时填写线路1和线路2名称",
                )
            line_names = (manual_line_1, manual_line_2)
            circuit_type = "双回"
        elif manual_line_1:
            line_names = (manual_line_1,)
            circuit_type = "单回"
        else:
            return ParsedLineIdentity(
                parsed_full_name,
                parsed_full_name,
                voltage,
                (),
                "",
                manual_required=True,
                manual_reason="手动输入缺少线路1名称",
            )
        if not voltage:
            return ParsedLineIdentity(
                parsed_full_name,
                parsed_full_name,
                "",
                line_names,
                circuit_type,
                manual_required=True,
                manual_reason="请补充电压等级",
            )
        return ParsedLineIdentity(parsed_full_name, parsed_full_name, voltage, line_names, circuit_type, tuple(warnings))

    if not voltage:
        warnings.append("未从文件名识别电压等级")

    double_match = re.search(r"(?P<base>.*?)(?:甲乙|甲、乙|甲/乙)线$", parsed_full_name)
    if double_match:
        base = double_match.group("base")
        line_names = ((base + "甲线") if base else "甲线", (base + "乙线") if base else "乙线")
        circuit_type = "双回"
    elif re.search(r"(?:甲|乙)线$", parsed_full_name):
        line_names = (parsed_full_name,)
        circuit_type = "单回"
    elif parsed_full_name:
        line_names = (parsed_full_name,)
        circuit_type = "单回"
    else:
        line_names = ()
        circuit_type = ""

    manual_required = not voltage or not line_names
    reason_parts = []
    if not voltage:
        reason_parts.append("无法识别电压等级")
    if not line_names:
        reason_parts.append("无法识别线路名称")
    return ParsedLineIdentity(
        parsed_full_name,
        parsed_full_name,
        voltage,
        line_names,
        circuit_type,
        tuple(warnings),
        manual_required=manual_required,
        manual_reason="；".join(reason_parts),
    )


def _first_child_text(element: ET.Element, name: str) -> str:
    for child in element.iter():
        if _local_name(child.tag) == name.lower():
            return "".join(child.itertext()).strip()
    return ""


def _strip_html(value: str) -> str:
    value = unescape(value or "")
    return re.sub(r"<[^>]*>", " ", value)


def _tower_number(name: str, description: str) -> Optional[int]:
    """Extract explicit positive tower numbers without mistaking voltage for one."""

    for text in (_strip_html(description), name):
        for pattern in _TAG_NUMBER_PATTERNS:
            match = pattern.search(text)
            if match:
                value = int(match.group(1))
                if value > 0:
                    return value
        match = _PURE_NUMBER.match(text)
        if match:
            value = int(match.group(1))
            if value > 0:
                return value
    return None


def _parse_coordinate(text: str, index: int, source_name: str) -> tuple[float, float]:
    tuples = [item for item in re.split(r"\s+", text.strip()) if item]
    if not tuples:
        raise KmlConversionError(
            "COORDINATE_INVALID",
            f"第 {index} 个杆塔点缺少 coordinates: {source_name or '未命名'}",
        )
    parts = [item.strip() for item in tuples[0].split(",")]
    if len(parts) < 2:
        raise KmlConversionError(
            "COORDINATE_INVALID",
            f"第 {index} 个杆塔点坐标格式无效: {source_name or '未命名'}",
        )
    try:
        longitude = float(parts[0])
        latitude = float(parts[1])
    except (TypeError, ValueError) as exc:
        raise KmlConversionError(
            "COORDINATE_INVALID",
            f"第 {index} 个杆塔点坐标不是数值: {source_name or '未命名'}",
        ) from exc
    if not (-180 <= longitude <= 180 and -90 <= latitude <= 90):
        raise KmlConversionError(
            "COORDINATE_INVALID",
            f"第 {index} 个杆塔点坐标越界: {longitude}, {latitude}",
        )
    return longitude, latitude


def _decode_kml(raw: bytes) -> str:
    if not raw:
        raise KmlConversionError("KML_DECODE_FAILED", "KML 文件为空")
    if raw.startswith(b"\xff\xfe"):
        return raw[2:].decode("utf-16le")
    if raw.startswith(b"\xfe\xff"):
        return raw[2:].decode("utf-16be")
    if raw.startswith(b"\xef\xbb\xbf"):
        raw = raw[3:]
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise KmlConversionError("KML_DECODE_FAILED", "KML 文件不是受支持的 UTF-8/UTF-16 编码") from exc


def parse_kml(raw: bytes, *, fallback_sequence: bool = True) -> tuple[list[TowerPoint], tuple[str, ...]]:
    """Parse Point Placemarks and return sorted tower points plus warnings."""

    text = _decode_kml(raw)
    try:
        root = ET.fromstring(text)
    except ET.ParseError as exc:
        raise KmlConversionError("KML_XML_INVALID", f"KML XML 解析失败: {exc}") from exc

    placemarks = [element for element in root.iter() if _local_name(element.tag) == "placemark"]
    if not placemarks:
        raise KmlConversionError("NO_TOWER_POINTS", "KML 中未找到 Placemark")

    points: list[tuple[Optional[int], float, float, str]] = []
    ignored_non_points = 0
    for index, placemark in enumerate(placemarks, start=1):
        point = next((child for child in placemark.iter() if _local_name(child.tag) == "point"), None)
        if point is None:
            ignored_non_points += 1
            continue
        name = _first_child_text(placemark, "name")
        description = _first_child_text(placemark, "description")
        coordinate_text = _first_child_text(point, "coordinates")
        longitude, latitude = _parse_coordinate(coordinate_text, index, name)
        points.append((_tower_number(name, description), longitude, latitude, name))

    if not points:
        raise KmlConversionError("NO_TOWER_POINTS", "KML 中未找到带有效 Point 坐标的杆塔")

    warnings: list[str] = []
    if ignored_non_points:
        warnings.append(f"已忽略 {ignored_non_points} 个非 Point 的 Placemark")

    used_numbers = {tower for tower, *_ in points if tower is not None}
    next_fallback = 1
    normalized: list[TowerPoint] = []
    for tower, longitude, latitude, name in points:
        used_fallback = tower is None
        if tower is None:
            if not fallback_sequence:
                raise KmlConversionError("TOWER_NUMBER_INVALID", f"无法识别杆塔编号: {name or '未命名'}")
            while next_fallback in used_numbers:
                next_fallback += 1
            tower = next_fallback
            used_numbers.add(tower)
            next_fallback += 1
            warnings.append(f"杆塔 {name or '未命名'} 未识别编号，已按顺序使用 {tower}")
        normalized.append(TowerPoint(tower, longitude, latitude, name, used_fallback))

    numbers = [item.tower for item in normalized]
    duplicates = sorted({number for number in numbers if numbers.count(number) > 1})
    if duplicates:
        raise KmlConversionError("TOWER_NUMBER_DUPLICATED", f"存在重复杆塔编号: {', '.join(map(str, duplicates))}")
    normalized.sort(key=lambda item: item.tower)
    return normalized, tuple(warnings)


def build_xlsx(
    line_name: str,
    towers: Iterable[TowerPoint],
    *,
    line_full_name: str = "",
    voltage_level: str = "",
    circuit_type: str = "单回",
) -> bytes:
    if not line_name.strip():
        raise KmlConversionError("INVALID_FILE_NAME", "无法从 KML 文件名得到线路名称")
    tower_list = list(towers)
    if not tower_list:
        raise KmlConversionError("NO_TOWER_POINTS", "没有可写入台账的杆塔")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "经纬度台账"
    sheet.freeze_panes = "A2"
    export_headers = HEADERS + METADATA_HEADERS
    sheet.append(list(export_headers))
    for tower in tower_list:
        sheet.append([
            line_name,
            tower.tower,
            tower.longitude,
            tower.latitude,
            line_full_name,
            voltage_level,
            circuit_type,
        ])

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in sheet["C"][1:] + sheet["D"][1:]:
        cell.number_format = "0.000000"
    for column, width in {"A": 22, "B": 14, "C": 14, "D": 14, "E": 24, "F": 14, "G": 12}.items():
        sheet.column_dimensions[column].width = width
    table = Table(displayName="TowerLedger", ref=f"A1:G{len(tower_list) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def convert_kml_bytes_batch(
    raw: bytes,
    filename: str,
    *,
    fallback_sequence: bool = True,
    overrides: Optional[dict[str, object]] = None,
) -> BatchConversionResult:
    filename = os.path.basename(filename or "")
    if not filename.lower().endswith(SUPPORTED_EXTENSION):
        raise KmlConversionError("INVALID_FILE_TYPE", "只支持 .kml 文件")
    if len(raw) > MAX_INPUT_BYTES:
        raise KmlConversionError("FILE_TOO_LARGE", f"KML 文件超过 {MAX_INPUT_BYTES // (1024 * 1024)} MB 限制")
    identity = parse_line_identity(filename, overrides)
    if not identity.source_line_name:
        raise KmlConversionError("INVALID_FILE_NAME", "KML 文件名不能为空")
    if identity.manual_required:
        return BatchConversionResult(
            source_file_name=filename,
            source_line_name=identity.source_line_name,
            line_full_name=identity.line_full_name,
            voltage_level=identity.voltage_level,
            circuit_type=identity.circuit_type,
            artifacts=(),
            warnings=identity.warnings,
            manual_required=True,
            manual_reason=identity.manual_reason,
        )
    towers, warnings = parse_kml(raw, fallback_sequence=fallback_sequence)
    combined_warnings = tuple(identity.warnings) + tuple(warnings)
    artifacts: list[LedgerArtifact] = []
    for line_name in identity.line_names:
        xlsx_bytes = build_xlsx(
            line_name,
            towers,
            line_full_name=identity.line_full_name,
            voltage_level=identity.voltage_level,
            circuit_type=identity.circuit_type,
        )
        artifacts.append(LedgerArtifact(
            file_name=f"{line_name}经纬度台账.xlsx",
            line_name=line_name,
            tower_count=len(towers),
            warnings=combined_warnings,
            xlsx_bytes=xlsx_bytes,
            sha256=sha256(xlsx_bytes).hexdigest(),
            line_full_name=identity.line_full_name,
            voltage_level=identity.voltage_level,
            circuit_type=identity.circuit_type,
        ))
    return BatchConversionResult(
        source_file_name=filename,
        source_line_name=identity.source_line_name,
        line_full_name=identity.line_full_name,
        voltage_level=identity.voltage_level,
        circuit_type=identity.circuit_type,
        artifacts=tuple(artifacts),
        warnings=combined_warnings,
    )


def convert_kml_bytes(raw: bytes, filename: str, *, fallback_sequence: bool = True) -> ConversionResult:
    """Backward-compatible single-ledger API.

    Use :func:`convert_kml_bytes_batch` for a possible 甲乙双回 result.
    """

    batch = convert_kml_bytes_batch(raw, filename, fallback_sequence=fallback_sequence)
    if batch.manual_required:
        raise KmlConversionError("MANUAL_INPUT_REQUIRED", batch.manual_reason)
    if len(batch.artifacts) != 1:
        raise KmlConversionError("MULTIPLE_OUTPUTS", "该 KML 会生成多本台账，请使用批量转换接口")
    artifact = batch.artifacts[0]
    return ConversionResult(
        line_name=artifact.line_name,
        file_name=artifact.file_name,
        tower_count=artifact.tower_count,
        warnings=artifact.warnings,
        xlsx_bytes=artifact.xlsx_bytes,
        sha256=artifact.sha256,
        line_full_name=artifact.line_full_name,
        voltage_level=artifact.voltage_level,
        circuit_type=artifact.circuit_type,
    )


def convert_kml_file(input_path: str | os.PathLike[str], output_path: str | os.PathLike[str]) -> ConversionResult:
    input_path = Path(input_path)
    result = convert_kml_bytes(input_path.read_bytes(), input_path.name)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(result.xlsx_bytes)
    return result


def validate_ledger_bytes(raw: bytes, *, expected_line_name: Optional[str] = None) -> dict[str, object]:
    """Validate the workbook before a desktop client writes it to the ledger library."""

    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise KmlConversionError("OUTPUT_FILE_INVALID", f"输出文件不是合法 XLSX: {exc}") from exc
    if "经纬度台账" not in workbook.sheetnames:
        raise KmlConversionError("OUTPUT_SCHEMA_INVALID", "输出文件缺少“经纬度台账”工作表")
    sheet = workbook["经纬度台账"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or tuple(rows[0][:4]) != HEADERS:
        raise KmlConversionError("OUTPUT_SCHEMA_INVALID", "输出台账列名不符合约定")
    data = [row[:4] for row in rows[1:] if any(value not in (None, "") for value in row[:4])]
    if not data:
        raise KmlConversionError("OUTPUT_SCHEMA_INVALID", "输出台账没有数据行")
    if expected_line_name and any(row[0] != expected_line_name for row in data):
        raise KmlConversionError("OUTPUT_SCHEMA_INVALID", "输出台账线路名称与输入不一致")
    return {"line_name": data[0][0], "tower_count": len(data), "headers": HEADERS}
