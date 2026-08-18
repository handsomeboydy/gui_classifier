import io
import unittest

from openpyxl import load_workbook

from kml_ledger_converter import (
    KmlConversionError,
    convert_kml_bytes,
    convert_kml_bytes_batch,
    parse_line_identity,
    parse_kml,
    validate_ledger_bytes,
)


def kml(*placemarks: str, encoding="utf-8") -> bytes:
    text = """<?xml version="1.0"?><kml xmlns="http://www.opengis.net/kml/2.2"><Document>{}</Document></kml>""".format("".join(placemarks))
    return text.encode(encoding)


def point(name, coordinates, description=""):
    return f"<Placemark><name>{name}</name><description>{description}</description><Point><coordinates>{coordinates}</coordinates></Point></Placemark>"


def line_string(name="线路边界"):
    return f"<Placemark><name>{name}</name><LineString><coordinates>113,23 114,24</coordinates></LineString></Placemark>"


class KmlLedgerConverterTests(unittest.TestCase):
    def test_parses_n_tower_names_and_double_line_outputs(self):
        raw = kml(
            point("220kV砚利甲乙线N2", "113.2515,23.500000,0"),
            point("220kV砚利甲乙线N1", "113.2500,23.500000,0"),
        )
        identity = parse_line_identity("220kV砚利甲乙线N1-N120.kml")
        self.assertEqual(identity.voltage_level, "220")
        self.assertEqual(identity.line_names, ("砚利甲线", "砚利乙线"))
        batch = convert_kml_bytes_batch(raw, "220kV砚利甲乙线N1-N120.kml")
        self.assertFalse(batch.manual_required)
        self.assertEqual([artifact.line_name for artifact in batch.artifacts], ["砚利甲线", "砚利乙线"])
        self.assertEqual([artifact.tower_count for artifact in batch.artifacts], [2, 2])
        workbook = load_workbook(io.BytesIO(batch.artifacts[0].xlsx_bytes), read_only=True, data_only=True)
        rows = list(workbook["经纬度台账"].iter_rows(values_only=True))
        self.assertEqual(rows[0][:7], ("线路名称", "杆塔编号", "经度", "纬度", "线路全称", "电压等级(kV)", "线路类型"))
        self.assertEqual(rows[1][0], "砚利甲线")
        self.assertEqual(rows[1][4:7], ("砚利甲乙线", "220", "双回"))

    def test_manual_overrides_resolve_missing_voltage(self):
        raw = kml(point("自定义N1", "113.25,23.5"))
        batch = convert_kml_bytes_batch(
            raw,
            "异常线路N1.kml",
            overrides={"voltage_level": "110", "line_name_1": "异常线", "circuit_type": "单回"},
        )
        self.assertFalse(batch.manual_required)
        self.assertEqual(batch.artifacts[0].line_name, "异常线")
        self.assertEqual(batch.artifacts[0].voltage_level, "110")

    def test_missing_voltage_requires_manual_input(self):
        result = convert_kml_bytes_batch(kml(point("线路N1", "113.25,23.5")), "线路N1.kml")
        self.assertTrue(result.manual_required)
        self.assertIn("电压", result.manual_reason)

    def test_converts_sorted_ledger_and_validates_output(self):
        raw = kml(
            point("塔#2", "113.2515,23.500000,0"),
            point("塔#1", "113.2500,23.500000,0"),
            line_string(),
        )
        result = convert_kml_bytes(raw, "220kV测试线路A.kml")
        self.assertEqual(result.file_name, "测试线路A经纬度台账.xlsx")
        self.assertEqual(result.tower_count, 2)
        self.assertEqual(result.warnings, ("已忽略 1 个非 Point 的 Placemark",))

        checked = validate_ledger_bytes(result.xlsx_bytes, expected_line_name="测试线路A")
        self.assertEqual(checked["tower_count"], 2)
        workbook = load_workbook(io.BytesIO(result.xlsx_bytes), read_only=True, data_only=True)
        rows = list(workbook["经纬度台账"].iter_rows(values_only=True))
        self.assertEqual(rows[0][:4], ("线路名称", "杆塔编号", "经度", "纬度"))
        self.assertEqual([row[1] for row in rows[1:]], [1, 2])

    def test_supports_utf16le_bom(self):
        raw = b"\xff\xfe" + kml(point("#1", "113.25,23.5"), encoding="utf-16le")
        towers, warnings = parse_kml(raw)
        self.assertEqual(towers[0].tower, 1)
        self.assertFalse(warnings)

    def test_missing_number_uses_non_conflicting_fallback_with_warning(self):
        towers, warnings = parse_kml(kml(point("电压500kV", "113.25,23.5"), point("#2", "113.251,23.5")))
        self.assertEqual([tower.tower for tower in towers], [1, 2])
        self.assertTrue(any("未识别编号" in warning for warning in warnings))

    def test_duplicate_number_is_rejected(self):
        with self.assertRaises(KmlConversionError) as context:
            parse_kml(kml(point("#1", "113.25,23.5"), point("#1", "113.251,23.5")))
        self.assertEqual(context.exception.code, "TOWER_NUMBER_DUPLICATED")

    def test_invalid_coordinate_is_rejected(self):
        with self.assertRaises(KmlConversionError) as context:
            parse_kml(kml(point("#1", "181,23.5")))
        self.assertEqual(context.exception.code, "COORDINATE_INVALID")

    def test_wrong_extension_is_rejected(self):
        with self.assertRaises(KmlConversionError) as context:
            convert_kml_bytes(kml(point("#1", "113.25,23.5")), "线路.txt")
        self.assertEqual(context.exception.code, "INVALID_FILE_TYPE")
